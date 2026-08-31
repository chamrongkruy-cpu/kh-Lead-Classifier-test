import streamlit as st
import pandas as pd
import numpy as np
import re
import io
import json
import math
from typing import Tuple, Dict, Any, List, Optional

# Attempt optional library imports with safe fallbacks
try:
    from rapidfuzz import fuzz
    HAS_RAPIDFUZZ = True
except ImportError:
    HAS_RAPIDFUZZ = False

try:
    from geopy.distance import geodesic
    HAS_GEOPY = True
except ImportError:
    HAS_GEOPY = False

try:
    from shapely.geometry import Point, Polygon
    from shapely.wkt import loads as load_wkt
    HAS_SHAPELY = True
except ImportError:
    HAS_SHAPELY = False

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# Safe secret retrieval to prevent KeyError on Streamlit Cloud deployment
APP_PASSWORD = st.secrets.get("APP_PASSWORD", None) if "APP_PASSWORD" in st.secrets else None

st.set_page_config(
    page_title="Sales Ops · Cambodia Lead Classifier",
    page_icon="🎯",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Custom CSS for foodpanda / Delivery Hero Cambodia branding
st.markdown("""
<style>
    .main-title {
        font-size: 2.2rem;
        font-weight: 800;
        color: #D70F64; /* foodpanda Pink */
        margin-bottom: 0.1rem;
    }
    .sub-title {
        font-size: 1.05rem;
        color: #4A4A4A;
        margin-bottom: 1.5rem;
    }
    .stButton>button[kind="primary"] {
        background-color: #D70F64 !important;
        border-color: #D70F64 !important;
        color: white !important;
        font-weight: 700 !important;
        border-radius: 8px !important;
        padding: 0.6rem 1.2rem !important;
    }
    .stButton>button[kind="primary"]:hover {
        background-color: #B50B52 !important;
        border-color: #B50B52 !important;
    }
    div[data-testid="stMetricValue"] {
        font-size: 1.8rem;
        font-weight: 700;
    }
</style>
""", unsafe_allow_html=True)

def normalize_cambodian_text(text: Any) -> str:
    """
    Normalizes Khmer script, Romanized Khmer transliterations, and English text.
    Strips noise words, legal entities, special symbols, and excess spaces.
    """
    if pd.isna(text) or text is None:
        return ""
    
    s = str(text).strip()
    if not s:
        return ""

    # Convert to lowercase
    s = s.lower()

    # Remove common business entity suffixes in Cambodia
    noise_patterns = [
        r'\bco\.?,?\s*ltd\.?\b',
        r'\bco\.?,?\s*ltd\b',
        r'\binc\.?\b',
        r'\bexpress\b',
        r'\bcambodia\b',
        r'\bphnom\s*penh\b',
        r'\bkhmer\b',
        r'\benterprise\b',
        r'\bgroup\b'
    ]
    for pattern in noise_patterns:
        s = re.sub(pattern, '', s, flags=re.IGNORECASE)

    # Clean punctuation except Khmer unicode characters (\u1780-\u17ff)
    s = re.sub(r'[^\w\s\u1780-\u17ff]', ' ', s)
    
    # Standardize consecutive whitespace
    s = re.sub(r'\s+', ' ', s).strip()
    return s

def haversine_distance(lat1: float, lon1: float, lat2: float, lon2: float) -> float:
    """
    Calculates geodesic distance in meters using Haversine formula (fallback if geopy unavailable).
    """
    R = 6371000.0  # Earth radius in meters
    phi1 = math.radians(lat1)
    phi2 = math.radians(lat2)
    delta_phi = math.radians(lat2 - lat1)
    delta_lambda = math.radians(lon2 - lon1)

    a = (math.sin(delta_phi / 2.0) ** 2 +
         math.cos(phi1) * math.cos(phi2) * math.sin(delta_lambda / 2.0) ** 2)
    c = 2.0 * math.atan2(math.sqrt(a), math.sqrt(1.0 - a))
    return R * c

def calculate_distance(lat1: float, lon1: float, lat2: float, lon2: float) -> float:
    """Calculates distance between coordinates in meters using Geopy or Haversine."""
    try:
        if HAS_GEOPY:
            return geodesic((lat1, lon1), (lat2, lon2)).meters
        return haversine_distance(lat1, lon1, lat2, lon2)
    except Exception:
        return float('inf')

def resolve_column(df: pd.DataFrame, possible_names: List[str]) -> Optional[str]:
    """Finds the first matching column name in df (case-insensitive)."""
    if df is None or df.empty:
        return None
    cols_lower = {c.lower().strip(): c for c in df.columns}
    for name in possible_names:
        nl = name.lower().strip()
        if nl in cols_lower:
            return cols_lower[nl]
    return None

def find_crm_matches(
    lead_row: pd.Series,
    crm_df: pd.DataFrame,
    lead_cols: Dict[str, str],
    crm_cols: Dict[str, str],
    radius_meters: float = 200.0,
    p4_threshold: float = 0.75,
    p3_threshold: float = 0.50
) -> Tuple[str, float, Optional[pd.Series], str]:
    """
    Executes Cambodian geographic and linguistic matching cascade:
    1. Lat/Lng GPS Proximity within radius_meters OR exact Sangkat/Khan match.
    2. Fuzzy Multilingual Name matching via RapidFuzz.
    """
    lead_name = normalize_cambodian_text(lead_row.get(lead_cols.get('name', '')))
    if not lead_name:
        return "P2 — Please Check", 0.0, None, "Blank Lead Name"

    lead_lat = lead_row.get(lead_cols.get('lat', ''))
    lead_lng = lead_row.get(lead_cols.get('lng', ''))
    lead_sangkat = normalize_cambodian_text(lead_row.get(lead_cols.get('sangkat', '')))

    has_coords = pd.notnull(lead_lat) and pd.notnull(lead_lng)
    try:
        if has_coords:
            lead_lat = float(lead_lat)
            lead_lng = float(lead_lng)
    except Exception:
        has_coords = False

    best_score = 0.0
    best_match_row = None
    match_reason = "No CRM Match"

    for _, crm_row in crm_df.iterrows():
        crm_lat = crm_row.get(crm_cols.get('lat', ''))
        crm_lng = crm_row.get(crm_cols.get('lng', ''))
        crm_sangkat = normalize_cambodian_text(crm_row.get(crm_cols.get('sangkat', '')))
        
        in_proximity = False
        dist_m = float('inf')

        # Proximity Check 1: GPS Radius
        if has_coords and pd.notnull(crm_lat) and pd.notnull(crm_lng):
            try:
                dist_m = calculate_distance(lead_lat, lead_lng, float(crm_lat), float(crm_lng))
                if dist_m <= radius_meters:
                    in_proximity = True
            except Exception:
                pass

        # Proximity Check 2: Sangkat / Khan Boundary Match
        if not in_proximity and lead_sangkat and crm_sangkat:
            if lead_sangkat in crm_sangkat or crm_sangkat in lead_sangkat:
                in_proximity = True

        if in_proximity:
            crm_name = normalize_cambodian_text(crm_row.get(crm_cols.get('name', '')))
            if not crm_name:
                continue

            if HAS_RAPIDFUZZ:
                score = max(
                    fuzz.token_set_ratio(lead_name, crm_name),
                    fuzz.token_sort_ratio(lead_name, crm_name)
                ) / 100.0
            else:
                score = 1.0 if lead_name == crm_name else (0.6 if lead_name in crm_name else 0.0)

            if score > best_score:
                best_score = score
                best_match_row = crm_row
                if dist_m < float('inf'):
                    match_reason = f"Proximity Match ({int(dist_m)}m)"
                else:
                    match_reason = f"Commune/Sangkat Match ({crm_sangkat})"

    if best_score >= p4_threshold:
        return "P4 — Duplicate", best_score * 100.0, best_match_row, match_reason
    elif best_score >= p3_threshold:
        return "P3 — Potential Match", best_score * 100.0, best_match_row, match_reason
    else:
        return "No CRM Match", 0.0, None, "No CRM match above threshold"

def validate_apify_status(
    apify_row: Optional[pd.Series],
    apify_cols: Dict[str, str],
    valid_categories: List[str]
) -> Tuple[str, str]:
    """Validates lead against Google Maps scrape output from Apify."""
    if apify_row is None:
        return "P2 — Please Check", "No Apify result found on Google Maps"

    cat_val = str(apify_row.get(apify_cols.get('category', ''), '')).lower().strip()
    perm_closed = str(apify_row.get(apify_cols.get('perm_closed', ''), '')).lower() in ['true', '1', 'yes']
    temp_closed = str(apify_row.get(apify_cols.get('temp_closed', ''), '')).lower() in ['true', '1', 'yes']

    if perm_closed or temp_closed:
        return "Business Closed", "Permanently or temporarily closed on Google Maps"

    if not cat_val:
        return "P2 — Please Check", "Category unmapped in Google Maps"

    is_eligible = any(cat in cat_val or cat_val in cat for cat in valid_categories)
    if is_eligible:
        return "P1 — New", f"Confirmed Open F&B venue ({cat_val})"
    else:
        return "Wrong Target Group", f"Non-F&B category ({cat_val})"

def export_to_excel(df_results: pd.DataFrame) -> bytes:
    """Generates a color-coded Excel workbook with 6 sheets for sales operations."""
    output = io.BytesIO()
    
    if not HAS_OPENPYXL:
        df_results.to_csv(output, index=False)
        return output.getvalue()

    wb = Workbook()
    
    fill_p1 = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Green
    fill_p4 = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Red
    fill_p3 = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # Yellow
    fill_p2 = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")  # Grey
    header_fill = PatternFill(start_color="D70F64", end_color="D70F64", fill_type="solid") # DH Pink
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

    ws_all = wb.active
    ws_all.title = "Classified Leads"

    headers = list(df_results.columns)
    ws_all.append(headers)
    for col_num in range(1, len(headers) + 1):
        cell = ws_all.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font

    for _, row in df_results.iterrows():
        ws_all.append(list(row))
        curr_row = ws_all.max_row
        lbl = str(row.get('Final Classification', ''))
        
        target_cell = ws_all.cell(row=curr_row, column=headers.index('Final Classification') + 1)
        if "P1" in lbl:
            target_cell.fill = fill_p1
        elif "P4" in lbl:
            target_cell.fill = fill_p4
        elif "P3" in lbl:
            target_cell.fill = fill_p3
        elif "P2" in lbl or "Closed" in lbl or "Wrong" in lbl:
            target_cell.fill = fill_p2

    sheets_config = [
        ("✅ P1 — New", df_results[df_results['Final Classification'] == "P1 — New"]),
        ("🔴 P4 — Duplicate", df_results[df_results['Final Classification'] == "P4 — Duplicate"]),
        ("🟡 P3 — Potential", df_results[df_results['Final Classification'] == "P3 — Potential Match"]),
        ("⚪ P2 — Please Check", df_results[df_results['Final Classification'] == "P2 — Please Check"]),
        ("⚠️ Closed & Wrong TG", df_results[df_results['Final Classification'].isin(["Business Closed", "Wrong Target Group"])])
    ]

    for title, sub_df in sheets_config:
        ws = wb.create_sheet(title=title)
        ws.append(headers)
        for col_num in range(1, len(headers) + 1):
            c = ws.cell(row=1, column=col_num)
            c.fill = header_fill
            c.font = header_font
        for _, r in sub_df.iterrows():
            ws.append(list(r))

    wb.save(output)
    return output.getvalue()

st.sidebar.title("⚙️ Settings")

MARKETS = {
    "Cambodia 🇰🇭 (KH)": "KH",
    "Singapore 🇸🇬 (SG)": "SG"
}

selected_market_label = st.sidebar.selectbox(
    "Market",
    options=list(MARKETS.keys()),
    index=0,
    help="Select the region to load local matching rules and boundary settings."
)

market_code = MARKETS[selected_market_label]

st.sidebar.subheader("Match Thresholds")

p3_threshold = st.sidebar.slider(
    "P3 Potential Match starts at (%)",
    min_value=40, max_value=80, value=50, step=5,
    help="Name similarity score range for manual rep verification."
) / 100.0

p4_threshold = st.sidebar.slider(
    "P4 Duplicate starts at (%)",
    min_value=60, max_value=95, value=75, step=5,
    help="Name similarity score at or above which a lead is marked as a Duplicate."
) / 100.0

proximity_radius = st.sidebar.slider(
    "GPS Proximity Radius (Meters)",
    min_value=50, max_value=1000, value=200, step=50,
    help="Radius around lead GPS coordinates to scan for existing CRM accounts."
)

default_categories = [
    "restaurant", "cafe", "coffee", "bakery", "food", "noodle", "asian restaurant",
    "fast food", "bubble tea", "dessert", "barbecue", "khmer restaurant", "bistro", "pub"
]

fnb_categories_input = st.sidebar.text_area(
    "Eligible F&B Categories (comma separated)",
    value=", ".join(default_categories),
    help="Google Maps categories considered valid for food delivery onboarding."
)

valid_categories_list = [c.strip().lower() for c in fnb_categories_input.split(",") if c.strip()]

st.markdown('<p class="main-title">🎯 Sales Ops · Cambodia Lead Classifier</p>', unsafe_allow_html=True)
st.markdown('<p class="sub-title"><b>Delivery Hero / foodpanda Cambodia</b> · Digital Sales APAC — Phnom Penh & Provinces</p>', unsafe_allow_html=True)

# Password Gate Check
if APP_PASSWORD:
    pwd_input = st.sidebar.text_input("🔑 App Password", type="password")
    if pwd_input != APP_PASSWORD:
        st.info("🔒 Please enter the correct password in the sidebar to access the classifier tool.")
        st.stop()

tab1, tab2, tab3, tab4 = st.tabs([
    "📊 Classify Leads",
    "🔗 Generate Apify URLs",
    "🏢 SF Account Audit",
    "📖 How to Use"
])

with tab1:
    # 📎 Collapsible guide to export required Salesforce reports
    with st.expander("📎 How to get your files — click to expand", expanded=False):
        c1, c2 = st.columns(2)
        with c1:
            st.markdown("**Step 1 · Leads file (Salesforce)**")
            st.link_button(
                "Open Cambodia Leads Report →",
                "https://deliveryhero.lightning.force.com/lightning/r/Report/00ObO000008clm5UAA/view?",
                use_container_width=True
            )
        with c2:
            st.markdown("**Step 2 · CRM Export (Salesforce)**")
            st.link_button(
                "Open Cambodia CRM Report →",
                "https://deliveryhero.lightning.force.com/lightning/r/Report/00ObO000008cmAHUAY/view?queryScope=userFolders",
                use_container_width=True
            )
        st.info(
            "🔗 **For Apify Results:** Go to the **Generate Apify URLs** tab → "
            "Step 1 generates your search URLs → paste into Apify Google Maps scraper → "
            "Step 2 enriches the output with the GRID column automatically."
        )

    st.subheader("1. Upload Input Files")
    st.caption("Upload the Salesforce Leads export, Apify Google Maps scrape results, and Salesforce All Accounts CRM export.")

    col1, col2, col3 = st.columns(3)
    
    with col1:
        leads_file = st.file_uploader("1️⃣ SF Leads File (.xlsx / .csv)", type=["xlsx", "csv"], key="leads")
    with col2:
        apify_file = st.file_uploader("2️⃣ Apify Results File (.xlsx / .csv)", type=["xlsx", "csv"], key="apify")
    with col3:
        crm_file = st.file_uploader("3️⃣ CRM All Accounts (.xlsx / .csv)", type=["xlsx", "csv"], key="crm")

    zones_file = st.file_uploader("🗺️ (Optional) Delivery Zones GeoJSON/JSON (`zones_KH.json`)", type=["json"], key="zones")

    loaded_zones = []
    if zones_file:
        try:
            loaded_zones = json.load(zones_file)
            st.success(f"Loaded {len(loaded_zones)} delivery coverage zones.")
        except Exception as e:
            st.error(f"Error loading zones JSON: {e}")

    if st.button("🚀 Run Lead Classification", type="primary", use_container_width=True):
        if not leads_file or not crm_file:
            st.error("Please upload at least the **Leads File** and the **CRM All Accounts File** to proceed.")
        else:
            with st.spinner("Processing files and executing Cambodian matching cascade..."):
                try:
                    df_leads = pd.read_excel(leads_file) if leads_file.name.endswith('.xlsx') else pd.read_csv(leads_file)
                    df_crm = pd.read_excel(crm_file) if crm_file.name.endswith('.xlsx') else pd.read_csv(crm_file)
                    
                    df_apify = None
                    if apify_file:
                        df_apify = pd.read_excel(apify_file) if apify_file.name.endswith('.xlsx') else pd.read_csv(apify_file)

                    lead_cols = {
                        'grid': resolve_column(df_leads, ['GRID', 'Lead ID', 'Id', 'Lead_GRID']),
                        'name': resolve_column(df_leads, ['Company / Account', 'Company', 'Lead Name', 'Account Name', 'Name']),
                        'street': resolve_column(df_leads, ['Street / Street No.', 'Street', 'Address', 'Street Address']),
                        'sangkat': resolve_column(df_leads, ['Sangkat / Khan / Province', 'Sangkat', 'District', 'City', 'State']),
                        'lat': resolve_column(df_leads, ['Coordinates (Latitude)', 'Latitude', 'Lat', 'location/lat']),
                        'lng': resolve_column(df_leads, ['Coordinates (Longitude)', 'Longitude', 'Lng', 'Lng/Lat', 'location/lng'])
                    }

                    crm_cols = {
                        'grid': resolve_column(df_crm, ['GRID', 'Account ID', 'Id']),
                        'name': resolve_column(df_crm, ['Account Name', 'Company Name', 'Name']),
                        'sangkat': resolve_column(df_crm, ['Sangkat / Khan', 'Sangkat', 'District', 'BillingCity']),
                        'lat': resolve_column(df_crm, ['Latitude', 'Lat', 'BillingLatitude']),
                        'lng': resolve_column(df_crm, ['Longitude', 'Lng', 'BillingLongitude']),
                        'address': resolve_column(df_crm, ['Formatted Restaurant Address', 'BillingStreet', 'Address'])
                    }

                    apify_cols = {}
                    if df_apify is not None:
                        apify_cols = {
                            'grid': resolve_column(df_apify, ['GRID', 'lead_grid', 'Input_GRID']),
                            'title': resolve_column(df_apify, ['title', 'name', 'placeName']),
                            'category': resolve_column(df_apify, ['categoryName', 'category', 'primaryCategory']),
                            'perm_closed': resolve_column(df_apify, ['permanentlyClosed', 'permanently_closed']),
                            'temp_closed': resolve_column(df_apify, ['temporarilyClosed', 'temporarily_closed'])
                        }

                    results = []

                    for idx, lead_row in df_leads.iterrows():
                        grid_val = lead_row.get(lead_cols.get('grid', ''), idx)
                        
                        # 1. CRM Match Cascade
                        crm_label, score, match_acc, match_reason = find_crm_matches(
                            lead_row, df_crm, lead_cols, crm_cols,
                            radius_meters=proximity_radius,
                            p4_threshold=p4_threshold,
                            p3_threshold=p3_threshold
                        )

                        final_label = crm_label
                        reason = match_reason
                        matched_crm_name = match_acc.get(crm_cols.get('name', ''), '') if match_acc is not None else ""
                        matched_crm_grid = match_acc.get(crm_cols.get('grid', ''), '') if match_acc is not None else ""

                        # 2. Apify Validation if no CRM match
                        if crm_label == "No CRM Match":
                            apify_match_row = None
                            if df_apify is not None and apify_cols.get('grid'):
                                matched_rows = df_apify[df_apify[apify_cols['grid']].astype(str) == str(grid_val)]
                                if not matched_rows.empty:
                                    apify_match_row = matched_rows.iloc[0]

                            apify_label, apify_reason = validate_apify_status(
                                apify_match_row, apify_cols, valid_categories_list
                            )
                            final_label = apify_label
                            reason = apify_reason

                        res_row = lead_row.to_dict()
                        res_row['Final Classification'] = final_label
                        res_row['Classification Reason'] = reason
                        res_row['Match Score (%)'] = round(score, 1)
                        res_row['Matched CRM Account Name'] = matched_crm_name
                        res_row['Matched CRM GRID'] = matched_crm_grid
                        
                        results.append(res_row)

                    out_df = pd.DataFrame(results)

                    st.divider()
                    st.subheader("2. Classification Results Summary")

                    m1, m2, m3, m4, m5 = st.columns(5)
                    m1.metric("Total Leads", len(out_df))
                    m2.metric("✅ P1 — New", len(out_df[out_df['Final Classification'] == "P1 — New"]))
                    m3.metric("🔴 P4 — Duplicate", len(out_df[out_df['Final Classification'] == "P4 — Duplicate"]))
                    m4.metric("🟡 P3 — Potential", len(out_df[out_df['Final Classification'] == "P3 — Potential Match"]))
                    m5.metric("⚪ P2 / Closed", len(out_df[out_df['Final Classification'].isin(["P2 — Please Check", "Business Closed", "Wrong Target Group"])]))

                    st.dataframe(out_df, use_container_width=True)

                    excel_bytes = export_to_excel(out_df)
                    st.download_button(
                        label="📥 Download Excel Report (Color-Coded Sheets)",
                        data=excel_bytes,
                        file_name="Cambodia_Classified_Leads_Report.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        type="primary"
                    )

                except Exception as e:
                    st.error(f"Error during classification: {str(e)}")
                    st.exception(e)

with tab2:
    st.subheader("🔗 Generate Apify Google Maps Search URLs")
    st.caption("Generate targeted Google Maps search URLs for Cambodian cities and provinces to feed into the Apify scraper.")

    raw_leads_text = st.text_area(
        "Paste Restaurant Names & Sangkats (one per line)",
        value="Bay Cha BKK1, Phnom Penh\nNum Banh Chok Toul Kork, Phnom Penh\nPub Street Cafe, Siem Reap",
        height=150
    )

    if st.button("Generate Search URLs"):
        lines = [line.strip() for line in raw_leads_text.split('\n') if line.strip()]
        url_data = []
        for line in lines:
            query = f"{line}, Cambodia"
            encoded_q = re.sub(r'\s+', '+', query)
            url = f"https://www.google.com/maps/search/{encoded_q}"
            url_data.append({"Search Query": line, "Google Maps Search URL": url})
        
        url_df = pd.DataFrame(url_data)
        st.dataframe(url_df, use_container_width=True)

with tab3:
    st.subheader("🏢 Salesforce CRM Internal Duplicate Audit")
    st.caption("Upload your Salesforce CRM All Accounts list to discover duplicate records already existing inside Salesforce.")

    crm_audit_file = st.file_uploader("Upload CRM Accounts File for Internal Audit", type=["xlsx", "csv"], key="crm_audit")

    if crm_audit_file and st.button("Run Internal Audit"):
        with st.spinner("Analyzing CRM records for duplicates..."):
            df_audit = pd.read_excel(crm_audit_file) if crm_audit_file.name.endswith('.xlsx') else pd.read_csv(crm_audit_file)
            name_col = resolve_column(df_audit, ['Account Name', 'Name', 'Company'])

            if not name_col:
                st.error("Could not locate 'Account Name' column in uploaded file.")
            elif not HAS_RAPIDFUZZ:
                st.error("Missing package 'rapidfuzz'. Please run `pip install rapidfuzz`.")
            else:
                duplicates = []
                num_records = min(len(df_audit), 400)
                records = df_audit.head(num_records).to_dict('records')

                for i in range(len(records)):
                    for j in range(i + 1, len(records)):
                        r1, r2 = records[i], records[j]
                        n1, n2 = normalize_cambodian_text(r1.get(name_col)), normalize_cambodian_text(r2.get(name_col))
                        
                        if n1 and n2:
                            score = fuzz.token_set_ratio(n1, n2) / 100.0
                            if score >= p4_threshold:
                                duplicates.append({
                                    "Account 1": r1.get(name_col),
                                    "Account 2": r2.get(name_col),
                                    "Similarity Score (%)": round(score * 100.0, 1),
                                    "Audit Action": "Flagged Internal Duplicate"
                                })

                if duplicates:
                    st.warning(f"Found {len(duplicates)} duplicate pairs within Salesforce CRM records!")
                    st.dataframe(pd.DataFrame(duplicates), use_container_width=True)
                else:
                    st.success("No internal duplicate records found above threshold.")

with tab4:
    st.markdown("""
    ### 📖 Cambodian Lead Classifier Guide
    
    #### Matching Logic Overview
    1. **Geographic Proximity First:** Evaluates CRM accounts within a configurable radius (default 200m) or within the same **Sangkat / Khan** (commune/district).
    2. **Multilingual Text Matching:** Normalizes Khmer script (`ភាសាខ្មែរ`), Romanized Khmer transliterations, and English characters.
    3. **Apify Google Maps Validation:** Verifies if non-CRM matched leads are active open food venues.

    #### Required Salesforce Export Fields
    - `Company / Account`
    - `Street / Street No.`
    - `Sangkat / Khan / Province`
    - `Coordinates (Latitude)` & `Coordinates (Longitude)`
    """)

